from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import datetime
import time

def get_today_sheet(workbook):
    """Returns the sheet matching today's day of the week."""
    days = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    today = datetime.datetime.now().strftime("%A")
    if today in days:
        return workbook[today]
    else:
        raise ValueError(f"No sheet found for the day: {today}")

def main():
    # Path to ChromeDriver
    driver_path = "path/to/chromedriver"
    driver = webdriver.Chrome(executable_path=driver_path)

    try:
        # Load Excel workbook
        workbook_path = "path/to/excel/file.xlsx"
        workbook = load_workbook(workbook_path)
        sheet = get_today_sheet(workbook)

        # Iterate through rows in the sheet
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):  # Assuming data starts from row 2
            keyword_cell = row[1]  # Assuming keyword is in column B
            if keyword_cell.value:
                keyword = keyword_cell.value

                # Perform Google search
                driver.get("https://www.google.com")
                search_box = driver.find_element(By.NAME, "q")
                search_box.send_keys(keyword)
                search_box.submit()

                # Wait for search results to load
                time.sleep(3)

                # Get search results
                results = driver.find_elements(By.CSS_SELECTOR, "h3")
                longest, shortest = "", ""

                for result in results:
                    text = result.text
                    if text:
                        if len(text) > len(longest):
                            longest = text
                        if not shortest or len(text) < len(shortest):
                            shortest = text

                # Write back to the sheet
                row[2].value = longest  # Longest option in column C
                row[3].value = shortest  # Shortest option in column D

        # Save the workbook
        workbook.save(workbook_path)

    except Exception as e:
        print("Error:", e)

    finally:
        # Close the browser
        driver.quit()

if _name_ == "_main_":
    main()
